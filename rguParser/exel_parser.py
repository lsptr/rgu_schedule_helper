import os
import openpyxl
from datetime import datetime
import sys
import subprocess
from db import db
from openpyxl.utils import range_boundaries


class ExcelParser:
    def __init__(self, parser_db):
        self.storage_path = "C:\\Users\\cepsk\\PycharmProjects\\RguScheduleHelper\\downloaded_files"
        self.parser_db = parser_db
        self.EXCLUDED_SHEETS = [
            "МАЙНОРЫ",
            "МАЙНОР",
            "ИТИЦ-2 курс",
            "ИТИЦ-3 курс",
            "29-3 курс",
            "29 -2 курс"
        ]


    def parse_all_excel_files(self):
        """Анализирует все Excel-файлы в папке"""
        k = 0
        if not os.path.exists(self.storage_path):
            print(f"Директория не найдена: {self.storage_path}")
            return

        for filename in os.listdir(self.storage_path):

            if filename.endswith(('.xlsx', '.xls')):
                filepath = os.path.join(self.storage_path, filename)

                if self.parser_db.is_schedule(filepath):
                    self.parse_excel_file(filename, filepath)
                    k += 1
        print("Tryed to parse " + str(k) + " files" )

    def parse_excel_file(self, filename, filepath):
        """Анализирует конкретный Excel-файл"""
        try:
            workbook = openpyxl.load_workbook(filepath, read_only=False, data_only=True)
            print(f"\nФайл: {filename}")

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                result = self.check_cell(sheet, 'B2', 'ФГБОУ ВО')
                print(f"  Лист: {sheet_name} - {'Y' if result else 'NO'}")

                if not result:
                    if not any(excluded.lower() in sheet_name.lower() for excluded in self.EXCLUDED_SHEETS):
                        self.parser_db.not_schedule(filepath)
                        print("__NOT SCHEDULE__")
                        # self.open_excel_file(filepath)
                        break

                data = self.extract_schedule_from_sheet(sheet)

                success = False
                if data:
                    success = self.parser_db.insert_data(data)
                else:
                    print("__NO DATA__")

                if success:
                    print("Данные успешно сохранены")
                    self.parser_db.save_parse_date(filepath)
                else:
                    print("Произошла ошибка при сохранении данных")

                if not self.parser_db.check_parse_date(filepath):
                    self.parser_db.not_schedule(filepath)

        except Exception as e:
            print(f"Ошибка при обработке файла {filename}: {e}")
        finally:
            if 'workbook' in locals():
                workbook.close()

    def extract_schedule_from_sheet(self, sheet):
        group_name = None

        for row in sheet.iter_rows(min_row=12, max_row=14):
            for cell in row:
                if cell.value and "ГРУППА" in str(cell.value):
                    group_name = str(cell.value).replace("ГРУППА", "").strip()
                    break
            if group_name:
                break

        if not group_name:
            print("НЕ НАЙДЕНА ГРУППА")
            return

        schedule_data = {
            "group_name": group_name,
            "lessons": []
        }

        header_row = None
        for row in sheet.iter_rows(min_row=13, max_row=15):
            for cell in row:
                if cell.value and "День недели" in str(cell.value):
                    header_row = row[0].row
                    break
            if header_row:
                break

        if not header_row:
            print("НЕ НАЙДЕН ЗАГОЛОВОК")
            return

        current_day = ""

        for row in sheet.iter_rows(min_row=header_row + 1, max_row=sheet.max_row):

            if row[1].value in ["ПН", "ВТ", "СР", "ЧТ", "ПТ", "СБ", "ВС"]:
                current_day = row[1].value

            if not current_day:
                continue

            if row[2].value:
                # day_of_week = current_day
                pair_number = str(row[2].value) if row[2].value else ""
                time = row[3].value if row[3].value else ""
                classroom_odd = str(row[4].value) if row[4].value else ""
                lesson_type_odd = row[5].value if row[5].value else ""
                teacher_odd = row[6].value if row[6].value else ""
                lesson_name_odd = row[7].value if row[7].value else ""
                lesson_name_even = row[8].value if row[8].value else ""
                teacher_even = row[9].value if row[9].value else ""
                lesson_type_even = row[10].value if row[10].value else ""
                classroom_even = str(row[11].value) if row[11].value else ""

                if (lesson_type_odd or lesson_type_even):

                    lesson_entry = {
                        "day_of_week": current_day,
                        "pair_number": pair_number,
                        "time": time,
                        "odd_week": {
                            "classroom": classroom_odd,
                            "lesson_type": lesson_type_odd,
                            "teacher": teacher_odd,
                            "lesson_name": lesson_name_odd
                        },
                        "even_week": {
                            "classroom": classroom_even,
                            "lesson_type": lesson_type_even,
                            "teacher": teacher_even,
                            "lesson_name": lesson_name_even
                        }
                    }
                    schedule_data["lessons"].append(lesson_entry)

        print(schedule_data)
        return schedule_data

    def check_cell(self, sheet, cell_coord, value):
        try:
            cell = sheet[cell_coord]

            # Получаем значение ячейки и приводим к строке
            cell_value = str(cell.value) if cell.value is not None else ""

            # Проверяем наличие искомого значения
            return value in cell_value

        except KeyError:
            print(f"Лист '{sheet}' не найден в файле {self.file_path}")
            return False
        except Exception as e:
            print(f"Ошибка при проверке ячейки {cell_coord}: {e}")
            return False

    def get_sheet_names(self):
        """Возвращает список всех листов в файле"""
        return self.workbook.sheetnames

    def get_cell_value(self, sheet, cell_coord):
        """Получает значение указанной ячейки"""
        try:
            return sheet[cell_coord].value
        except Exception as e:
            print(f"Ошибка при получении значения ячейки {cell_coord}: {e}")
            return None

    def open_excel_file(self, filepath):
        """Открывает Excel-файл в ассоциированной программе"""
        try:
            if os.name == 'nt':  # Windows
                os.startfile(filepath)
            elif os.name == 'posix':  # macOS/Linux
                subprocess.run(['open', filepath] if sys.platform == 'darwin' else ['xdg-open', filepath])
            print(f"Файл открыт для проверки: {filepath}")
        except Exception as e:
            print(f"Не удалось открыть файл {filepath}: {e}")
